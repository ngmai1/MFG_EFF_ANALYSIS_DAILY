import os
from pathlib import Path
import numpy as np
import mysql.connector
from datetime import date, datetime, timedelta
from sqlalchemy import create_engine
import pandas as pd
import openpyxl, xlrd
import pyodbc
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

hostname = 'pbvweb01v'
engine = create_engine('mysql+mysqlconnector://LaDat:l4d4t5@pbvweb01v:3306/pr2k', echo=False)
engine2 = create_engine('mysql+mysqlconnector://LaDat:l4d4t5@pbvpweb01:3306/pr2k', echo=False)
engineNam = create_engine('mysql+mysqlconnector://LaDat:l4d4t5@pbvweb01v:3306/erpsystem', echo=False)
engineLine = create_engine('mysql+mysqlconnector://LaDat:l4d4t5@pbvweb01v:3306/linebalancing', echo=False)
# ================================================offline=======================================================
wb_template=openpyxl.load_workbook('D:\\report\\Daily_eff_report\\DailyEfficiencyReportTemplate72.xlsx')
pr_list = ["PR", "IE", "HR", "AMTPR", "QA", "MEC", "LP", "IT", "FI", "PL", "LOG", "F&M", "CPL", "AD"]
cut_list = ["CT", "CTIE"]


def get_date_format(date):
    year = date[0:4]
    month = date[5:7]
    day = date[8:10]
    return year + month + day


def get_this_week(yesterday):
    week_str = ''
    week = int(yesterday.strftime("%W"))
    if week < 10:
        week_str = 'W0' + str(week)
    else:
        week_str = 'W' + str(week)
    return week_str


def findFile(yesterday, pathTimeSheet):
    filePath = ''
    for r, d, f in os.walk(pathTimeSheet):
        for entry in f:
            if yesterday in entry:
                if '~' not in entry:
                    filePath = os.path.join(r, entry)
    return filePath


def get_timesheet_by_date(date):
    connection = pyodbc.connect(
        # Trusted_Connection='Yes',
        DRIVER='{SQL Server}',
        SERVER='PBVPAYQSQL1V',
        UID='cts',
        PWD='Ct$yS123',
        database='PBCTS')

    sql = (
                "SELECT EMPLOYEE, DATE, DEPT, EMP_TYPE, SHIFT, TIME_IN, TIME_OUT, AbsentHrs, SOON, LATE, totalWorkedH, OT15, OT20, OT30, OTActual, totalWorkedOTH, LeaveCode, ShiftCode "
                + "FROM dbo.IT_TimeSheetInnovation o WHERE o.DATE = '" + date + "';")

    # sql=("select EMPLOYEE, DATE, DEPT, EMP_TYPE, SHIFT, TIME_IN, TIME_OUT, AbsentHrs, SOON, LATE, totalWorkedH, OT15, OT20, OT30, OTActual, totalWorkedOTH, le.LeaveCode, ShiftCode from "
    #     +" (select EmpID as EMPLOYEE, TSDATE as DATE, ts.DepartmentCode as DEPT, EmployeeTypeCode as EMP_TYPE, GroupCode as SHIFT, "
    #     +" TimeIn as TIME_IN, TimeOut as TIME_OUT, AbsentHrs, SOON, LATE, totalWorkedH, (OT15+OT15N) OT15, (OT20+OT20N) OT20, OT30, OTActual, totalWorkedOTH, ts.ShiftCode "
    #     +" from TMS_TimesheetDetails ts inner join [Employees Details] em on ts.EmpID=em.[Employee ID] where TSDate='"+date+"') t1 left join TMS_Leaves le on t1.EMPLOYEE=le.EmpID  and  t1.DATE=le.FromDate;")
    timesheet = pd.read_sql(sql, connection)
    connection.close()
    return timesheet


def delete_old_timesheet(date):
    mydb = mysql.connector.connect(host='pbvweb01v', user='LaDat', passwd='l4d4t5', database="pr2k")
    myCursor = mydb.cursor()
    query = ('delete from employee_timesheet where DATE="' + date + '"')
    myCursor.execute(query)
    mydb.commit()
    mydb.close()


def get_cd15_group(date):
    query = ("SELECT e.ID, t2.*, l.Location FROM (SELECT NameGroup, AVG(OffTime) OffTime FROM "
             " (SELECT t.ID, e.Line, l.NameGroup, Code, SUM(SpanTime) as OffTime "
             " from linebalancing.operation_offstandard_tracking t LEFT JOIN erpsystem.setup_emplist e ON t.ID=e.ID "
             " LEFT JOIN erpsystem.setup_location l ON e.Line=l.Location "
             " where DateUpdate='" + date + " 00:00:00' AND CODE LIKE '15%' and IEApprovedResult='2' "
                                            " GROUP BY t.ID, e.Line, l.NameGroup, LEFT(Code, 2)) t1 GROUP BY NameGroup) t2 LEFT JOIN erpsystem.setup_location l ON t2.NameGroup=l.NameGroup "
                                            " LEFT JOIN erpsystem.setup_emplist e ON l.Location=e.Line;")
    data = pd.read_sql(query, engine)
    engine.dispose()
    return data


def get_reg_hrs(time_in, time_out):
    timei = datetime.strptime(time_in, "%H:%M:%S")
    timeo = datetime.strptime(time_out, "%H:%M:%S")
    if timei != timeo:
        diff = timeo - timei
        hours = (diff.total_seconds()) / 3600
        reg_hrs = hours
        # if hours > 8:
        #     reg_hrs = 8
        # else:
        #     reg_hrs = hours
    else:
        reg_hrs = 8
    return reg_hrs


def process_timesheet(date, rit_whr, bal_whr, rit_ot, bal_ot):
    # =====Làm việc bình thường 8h====
    # -30 phút ăn cơm
    # -8 phút chất lượng
    # =====Làm việc tăng ca===========
    # Trên 2h:
    # -30 phút ăn cơm:
    # -4 phút chất lượng
    # =====Lay cd 15==================
    cd15_group = get_cd15_group(date)
    timesheet = []
    workHrsData = get_timesheet_by_date(date)
    for row in range(0, len(workHrsData)):
        ID = str(workHrsData.iloc[row, 0])
        if len(ID) == 5:
            ID = '0' + ID
        id5 = ID[1:6]
        datets = date
        dept = str(workHrsData.iloc[row, 2])
        typeEmp = str(workHrsData.iloc[row, 3])
        groupcd = str(workHrsData.iloc[row, 4])
        if 'n' in groupcd:
            groupcd = ''
        timein = str(workHrsData.iloc[row, 5])
        if 'n' in timein:
            timein = ''
        timeout = str(workHrsData.iloc[row, 6])
        if 'n' in timeout:
            timeout = ''
        if timein == '00:00:00' and timeout == '00:00:00':
            continue
        absent = str(workHrsData.iloc[row, 7])
        if 'n' in absent:
            absent = '0'
        soon = str(workHrsData.iloc[row, 8])
        if 'n' in soon:
            soon = '0'
        late = str(workHrsData.iloc[row, 9])
        if 'n' in late:
            late = '0'
        reg = str(workHrsData.iloc[row, 10])
        if 'n' in reg:
            reg = '0'
        otActual = str(workHrsData.iloc[row, 14])
        if 'n' in otActual:
            otActual = '0'
        ot15 = str(workHrsData.iloc[row, 11])
        if 'n' in ot15:
            ot15 = '0'
        ot20 = str(workHrsData.iloc[row, 12])
        if 'n' in ot20:
            ot20 = '0'
        ot30 = str(workHrsData.iloc[row, 13])
        if 'n' in ot30:
            ot30 = '0'
        otActual = str(workHrsData.iloc[row, 14])
        totalREG = str(workHrsData.iloc[row, 15])
        totalREGAdm2 = workHrsData.iloc[row, 15]
        if 'n' in totalREG:
            totalREG = '0'
        # g='0'
        shiftOT = 'N'
        if dept in pr_list:
            workActual = 0
            if 'RIT' in groupcd:
                workActual = rit_whr - float(soon) - float(late) - float(absent)
                if rit_whr == 0:
                    reg = '0'
            if 'BAL' in groupcd:
                workActual = bal_whr - float(soon) - float(late) - float(absent)
                if bal_whr == 0:
                    reg = '0'
            if workActual < 0:
                workActual = 0
                reg = '0'
            totalREG_raw = str(float(reg) + float(otActual))
            if typeEmp == 'DR':
                totalREG = str(float(workActual) + float(otActual))
            else:
                totalREG = totalREG_raw
            if 'RIT' in groupcd and rit_ot == "Y":
                totalREG = str(float(workActual))
                totalREG_raw = str(float(reg))
                reg = '0'
                shiftOT = 'Y'
            if 'BAL' in groupcd and bal_ot == "Y":
                totalREG = str(float(workActual))
                totalREG_raw = str(float(reg))
                reg = '0'
                shiftOT = 'Y'
            if 'RIT' in groupcd and rit_ot == "C":
                # reg=otActual
                otActual = '0'
                ot15 = '0'
                ot20 = '0'
                ot30 = '0'
            if 'BAL' in groupcd and bal_ot == "C":
                # reg=otActual
                otActual = '0'
                ot15 = '0'
                ot20 = '0'
                ot30 = '0'
            if float(totalREG) < 0:
                totalREG = '0'
            if float(totalREG_raw) < 0:
                totalREG_raw = '0'
        alcode = str(workHrsData.iloc[row, 16])
        if 'n' in alcode:
            alcode = ''
        shiftCode = str(workHrsData.iloc[row, 17])
        workHrs = float(totalREG)
        if dept in cut_list:
            if workHrs >= 10:
                workHrs = round(workHrs - 1 - 13 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
            elif workHrs >= 4:
                workHrs = round(workHrs - 0.5 - 9 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
            else:
                workHrs = round(workHrs - 4 / 60, 2)  # - 8 phút chất lượng
            if 'NS' in shiftCode and workHrs > 6:
                workHrs = workHrs - 0.5
        if (
                'RIT' == groupcd or 'RITS' == groupcd or 'RITQ' == groupcd or 'RITC-OT' in groupcd or 'RITM7-OT' in groupcd) and typeEmp == 'DR':
            if 'RITM7-OT' in groupcd:
                workHrs = workHrs - 1  # -float(otActual)
                if float(reg) > 7:
                    otActual = str(float(otActual) + (float(reg) - 1))
                    reg = '7'

                # reg=str(round(float(reg)-float(otActual),2))
            if rit_ot != 'Y':  # OT normal: OT<>Reg
                if workHrs >= 10:
                    workHrs = round(workHrs - 1 - 12 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
                elif workHrs >= 4:
                    workHrs = round(workHrs - 0.5 - 8 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
                else:
                    workHrs = round(workHrs - 4 / 60, 2)  # - 8 phút chất lượng
            else:  # OT=REG
                if workHrs >= 12:
                    workHrs = round(workHrs - 1 - 12 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
                elif workHrs >= 4:
                    workHrs = round(workHrs - 0.5 - 8 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
                else:
                    workHrs = round(workHrs - 4 / 60, 2)  # - 8 phút chất lượng
        if (
                'BALI' == groupcd or 'BALIS' == groupcd or 'BALIQ' == groupcd or 'BALIC-OT' in groupcd or 'BALIM7-OT' in groupcd) and typeEmp == 'DR':
            if 'BALIM7-OT' in groupcd:
                workHrs = workHrs - 1  # float(otActual)
                # reg=str(round(float(reg)-float(otActual),2))
                if float(reg) > 7:
                    otActual = str(float(otActual) + (float(reg) - 1))
                    reg = '7'
            if bal_ot != 'Y':  # OT normal: OT<>Reg
                if workHrs >= 10:
                    workHrs = round(workHrs - 1 - 12 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
                elif workHrs >= 4:
                    workHrs = round(workHrs - 0.5 - 8 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
                else:
                    workHrs = round(workHrs - 4 / 60, 2)  # - 8 phút chất lượng
            else:  # OT=REG
                if workHrs >= 4:
                    workHrs = round(workHrs - 0.5 - 8 / 60, 2)  # - ăn cơm, - 8 phút chất lượng
                else:
                    workHrs = round(workHrs - 4 / 60, 2)  # - 8 phút chất lượng
        # ///////////
        if 'MAT' in groupcd and 'RIT' in groupcd and typeEmp == 'DR':
            if rit_ot != 'Y':  # OT normal: OT<>Reg
                if workHrs > 8:
                    workHrs = 8
                if workHrs >= 4:
                    workHrs = round(workHrs - 1.5 - 8 / 60, 2)  # - ăn cơm, - 8 phút chất lượng NOTE=======1.5
                else:
                    workHrs = round(workHrs - 8 / 60, 2)  # - 4 phút chất lượng
            else:  # OT=REG
                workHrs = 0

        if 'CHIL' in groupcd and 'RIT' in groupcd and typeEmp == 'DR':
            if rit_ot != 'Y':  # OT normal: OT<>Reg
                if workHrs > 8:
                    workHrs = 8
                if workHrs >= 4:
                    workHrs = round(workHrs - 2.5 - 8 / 60, 2)  # - ăn cơm, - 8 phút chất lượng NOTE=======1.5
                else:
                    workHrs = round(workHrs - 8 / 60, 2)  # - 4 phút chất lượng
            else:  # OT=REG
                workHrs = 0

        if 'MAT' in groupcd and 'BAL' in groupcd and typeEmp == 'DR':
            if rit_ot != 'Y':  # OT normal: OT<>Reg
                if workHrs >= 8:
                    workHrs = 8  # - ăn cơm, - 8 phút chất lượng
                if workHrs >= 4:
                    workHrs = round(workHrs - 1.5 - 8 / 60, 2)  # - ăn cơm, - 8 phút chất lượng NOTE========1.5
                else:
                    workHrs = round(workHrs - 8 / 60, 2)  # - 8 phút chất lượng
            else:  # OT=REG
                workHrs = 0

        if 'CHIL' in groupcd and 'BAL' in groupcd and typeEmp == 'DR':
            if rit_ot != 'Y':  # OT normal: OT<>Reg
                if workHrs >= 8:
                    workHrs = 8  # - ăn cơm, - 8 phút chất lượng
                if workHrs >= 4:
                    workHrs = round(workHrs - 2.5 - 8 / 60, 2)  # - ăn cơm, - 8 phút chất lượng NOTE========1.5
                else:
                    workHrs = round(workHrs - 8 / 60, 2)  # - 8 phút chất lượng
            else:  # OT=REG
                workHrs = 0
        # //////////
        if 'C6-OT' in groupcd and typeEmp == 'DR':
            reg_hrs_c6ot = get_reg_hrs(workHrsData.iloc[row, 5], workHrsData.iloc[row, 6])
            workHrs = round(reg_hrs_c6ot - 0.5 - 8 / 60, 2)
        if 'ADM2' in groupcd and typeEmp == 'DR':
            reg_hrs_admin2 = get_reg_hrs(workHrsData.iloc[row, 5], workHrsData.iloc[row, 6])
            workHrs = round(totalREGAdm2 - 0.5 - 8 / 60, 2)
        if 'ADM2-C6' in groupcd and typeEmp == 'DR':
            reg_hrs_admin2 = get_reg_hrs(workHrsData.iloc[row, 5], workHrsData.iloc[row, 6])
            workHrs = round(totalREGAdm2 - 2.5 - 8 / 60, 2)
        if 'AD2-MAT' in groupcd and typeEmp == 'DR':
            reg_hrs_admin2 = get_reg_hrs(workHrsData.iloc[row, 5], workHrsData.iloc[row, 6])
            workHrs = round(totalREGAdm2 - 1.5 - 8 / 60, 2)
        if 'AD2-M7OT' in groupcd and typeEmp == 'DR':
            if float(reg) == 8:
                totalREG_raw = 7 + float(otActual)
            else:
                totalREG_raw = float(reg) + float(otActual)
            workHrs = round(totalREG_raw - 0.5 - 8 / 60, 2)
        if 'CHILD7' in groupcd and typeEmp == 'DR':
            reg_hrs_child7 = get_reg_hrs(workHrsData.iloc[row, 5], workHrsData.iloc[row, 6])
            workHrs = round(reg_hrs_child7 - 0.5 - 8 / 60, 2)
        if 'M7-OT' in groupcd and typeEmp == 'DR':
            # reg_hrs_m7ot = get_reg_hrs(workHrsData.iloc[row, 5], workHrsData.iloc[row, 6])
            # workHrs = round(reg_hrs_m7ot - 0.5 - 8 / 60, 2)
            workHrs = round(float(otActual) - 0.5 - 8 / 60, 2)
            if workHrs <= 1:
                reg_hrs_child7 = get_reg_hrs(workHrsData.iloc[row, 5], workHrsData.iloc[row, 6])
                workHrs = round(reg_hrs_child7 - 0.5 - 8 / 60, 2)
            if float(otActual) == 0.0 and float(totalREGAdm2) == 0.0:
                workHrs = 0

        if workHrs < 0:
            workHrs = 0
        # except:
        #     workHrs = 0
        cd03 = 0
        cd08 = 0
        cd09 = 0
        cd15 = 0
        if typeEmp == 'DR':
            dateFull = date[0:4] + '-' + date[4:6] + '-' + date[6:8]
            cd03, cd08, cd09 = get_offstd_time(str(int(ID)), groupcd, date)
            if dept == 'CT':
                cd03, cd08, cd09 = get_offstd_time_cutting(ID, groupcd, dateFull)
        IDkey = id5 + datets
        cd15 = cd15_group.query('ID=="' + ID + '"')['OffTime'].sum()
        if cd15 > 0:
            print(ID, cd15)
        thisTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # ===threshold reg for Mat and Child================
        if ('MAT' in groupcd or 'CHI' in groupcd) and float(reg) >= 7:
            reg = '8'
        timesheet.append({'ID': IDkey, 'ID5': id5, 'EMPLOYEE': ID, 'DEPT': dept, 'EMP_TYPE': typeEmp, 'SHIFT': groupcd,
                          'DATE': datets, 'TIME_IN': timein, 'TIME_OUT': timeout,
                          'ABSENT': absent, 'SOON': soon, 'LATE': late, 'REG_HRS': reg, 'OT15': ot15, 'OT20': ot20,
                          'OT30': ot30, 'OT_ACTUAL': otActual,
                          'WORK_HRS': workHrs, 'REG_HRS_TOTAL': totalREG_raw, 'CD03': cd03, 'CD08': cd08, 'CD09': cd09,
                          'CD15': cd15, 'AL_CODE': alcode, 'TimeUpdate': thisTime, 'OTCN': shiftOT})
    timesheetDF = pd.DataFrame(timesheet)
    delete_old_timesheet(date)
    timesheetDF.to_sql('employee_timesheet', con=engine, if_exists='append', index=False)
    engine.dispose()


def timesheet_existed(date):
    data = pd.read_sql('select ID from employee_timesheet where DATE="' + date + '"', engine)
    engine.dispose()
    if len(data) > 0:
        return True
    else:
        return False


def get_sum_offStd(date, employee):
    data = pd.read_sql('SELECT SUM(sp) AS sum_offstandard FROM (SELECT if(SpanTime IS NULL, 0, SpanTime) AS sp '
                       + 'FROM operation_offstandard_tracking '
                       + 'WHERE  DateUpdate="' + date + ' 00:00:00" AND left(CODE,2) '
                       + 'IN("02","04","06","07","10","11","12","13") AND IEApprovedResult = 2 AND right(ID,5) = right("' + employee + '",5)) b1;',
                       engineLine)
    engine.dispose()
    res = data.iloc[0, 0]
    if res == None:
        return 0
    else:
        return data.iloc[0, 0]


def get_sum_offStdc10(date, employee):
    data = pd.read_sql('SELECT SUM(sp) AS sum_offstandard FROM (SELECT if(SpanTime IS NULL, 0, SpanTime) AS sp '
                       + 'FROM operation_offstandard_tracking '
                       + 'WHERE  DateUpdate="' + date + ' 00:00:00" AND left(CODE,2) '
                       + 'IN("10") AND IEApprovedResult = 2 AND right(ID,5) = right("' + employee + '",5)) b1;',
                       engineLine)
    engine.dispose()
    res = data.iloc[0, 0]
    if res == None:
        return 0
    else:
        return data.iloc[0, 0]


def get_dkn(empQuery):
    data_sort = empQuery.sort_values(by=['SAH'], ascending=False)
    if len(empQuery) == 1:
        operation2 = ''
    else:
        operation2 = data_sort.iloc[1, 3]
    return operation2


def getEff(date):
    # pathTimeSheet='\\\\pbvfps1\\PBSHARE\\PR\\TIMESHEET\\2020'
    dateFull = date[0:4] + '-' + date[4:6] + '-' + date[6:8]
    workHrsData = pd.read_sql(
        'select EMPLOYEE, SHIFT, OT15, OT20, OT30, REG_HRS, OT_ACTUAL, (CD09+CD15) CD09, WORK_HRS, OTCN from employee_timesheet where DATE="' + dateFull + '" and EMP_TYPE="DR";',
        engine)
    engine.dispose()
    sql = ('select ID, Name, Line, Shift, Type, Dept from setup_emplist where Type="DR"')
    emplistData = pd.read_sql(sql, engineNam)
    engineNam.dispose()
    emplist_size = len(emplistData)
    row = 0
    dailyBundleData = pd.DataFrame()

    # fix loi cd null
    sql = ('select DATE, EMPLOYEE, OPERATION_CODE, OPERATION, COUNT(TICKET) as BUNDLE, SUM(EARNED_HOURS)/60 as '
           + ' SAH from employee_scanticket where DATE="' + date + '" AND operation IS NOT NULL '
           + ' group by EMPLOYEE, OPERATION_CODE;')
    daily_sew = pd.read_sql(sql, engine)
    sql = (
            'SELECT DATE_FORMAT(DATE, "%Y%m%d") as DATE, RIGHT(IDEmployees,5) as EMPLOYEE, ZoneMover as OPERATION_CODE, CONCAT("MOVER", ZoneMover) as OPERATION, SUM(DzCase) as BUNDLE, SUM(DzCase)*SAH AS SAH '
            + ' FROM data_finishedgoodssewing INNER JOIN setup_sahmover ON data_finishedgoodssewing.ZoneMover=setup_sahmover.Area '
            + ' WHERE DATE=DATE_FORMAT("' + date + '", "%Y-%m-%d") GROUP BY IDEmployees, ZoneMover;')
    daily_mover = pd.read_sql(sql, engineNam)
    dailyBundleData = dailyBundleData.append(daily_sew)
    dailyBundleData = dailyBundleData.append(daily_mover)
    sql = "SELECT EMPLOYEE, OPERATION FROM linebalancing.data_operation_7000_register"
    sah7000 = pd.read_sql(sql, engine)
    engine.dispose()
    engineNam.dispose()
    dataset = []
    dataset2 = []
    effDataset = []
    for row in range(0, emplist_size):
        ID = emplistData.iloc[row, 0]
        name = emplistData.iloc[row, 1]

        c02, c04, c05, c06, c07, c10, c11, c12, c13 = get_offStd_all(dateFull, ID)

        line = emplistData.iloc[row, 2]
        groupData = pd.read_sql('select NameGroup from setup_location where Location="' + line + '"', engineNam)
        if len(groupData) > 0:
            group = groupData.iloc[0, 0]
        else:
            group = '000-000'
        shift = emplistData.iloc[row, 3]
        employee_infor = workHrsData.query('EMPLOYEE=="' + ID + '"')
        workHrs = 0
        offStd = 0
        offStdc10 = 0
        otActual = 0
        cnOT = 'N'

        if len(employee_infor) > 0:
            workHrs = float(employee_infor.iloc[0, 8])
            ot15 = float(employee_infor.iloc[0, 2])
            ot20 = float(employee_infor.iloc[0, 3])
            ot30 = float(employee_infor.iloc[0, 4])
            otActual = float(employee_infor.iloc[0, 6])
            tem_ot = ''
            eff_ot = ''
            if otActual >= 2 and otActual <= 4:
                data = get_tem_ot(ID, dateFull)
                len_tem_ot = len(data)
                if len_tem_ot != 0:
                    tem_ot = len_tem_ot
                    sah = (data['earned_hours'].sum())/60
                    eff_ot = round((sah/otActual)*100, 2)
                else:
                    eff_ot = ''
                    tem_ot = ''

            cnOT = employee_infor.iloc[0, 9]
            plantEff = 0
            dolEff = 0
            offStd = float(employee_infor.iloc[0, 7])
            offStdc10 = float(get_sum_offStdc10(dateFull, ID))
            offStd2 = float(get_sum_offStd(dateFull, ID))
            regPayroll = float(employee_infor.iloc[0, 5])
        else:  # absent
            continue
        empQuery = dailyBundleData.query('EMPLOYEE=="' + str(ID[1:6]) + '"')
        code = ''
        op = ''
        bundle = 0
        sumbundle = 0
        sah = 0
        sumSAH = 0

        main_op = ''
        max_sah = 0
        is_sah7000 = ''
        dkn = ''
        machine_count = ''
        if len(empQuery) > 0:
            # lấy công đoạn đa kỹ năng
            if len(empQuery) > 1:
                data_sort = empQuery.sort_values(by=['SAH'], ascending=False)
                dkn = str(data_sort.iloc[1, 3])
                if dkn == 'None':
                    dkn = ''

                mydb = mysql.connector.connect(host='pbvweb01v', user='LaDat', passwd='l4d4t5', database="mms")
                mycursor = mydb.cursor()
                list = empQuery['OPERATION'].values.tolist()
                sql = f"SELECT COUNT(tb.machine_name) AS count FROM(SELECT * FROM mms.machine_operation m WHERE m.operation_name IN ({','.join(['%s'] * len(list))}) GROUP BY m.machine_name) tb;"
                mycursor.execute(sql, tuple(list))
                myresult = mycursor.fetchone()
                if myresult[0] == 0:
                    machine_count = ''
                else:
                    machine_count = str(myresult[0])

            # empQuery.sort_values(['SAH'], ascending=False).reset_index(drop=True)
            sumSAH = float(empQuery['SAH'].sum())
            if workHrs > 0:
                eff = 0
                if workHrs - (offStd2 + offStd) <= 0:
                    eff = 0
                else:
                    if workHrs - (offStd2 + offStd) < 0:
                        offStd2 = 0
                    eff = round(sumSAH / (workHrs - (offStd2 + offStd)) * 100, 2)
                plantEff = eff
                dolEff = round(sumSAH / workHrs * 100, 2)
            else:
                workHrs = 0
                eff = 0
                plantEff = 0
                dolEff = 0
            if workHrs <= 1:
                eff = ''
            workHrs = str(workHrs)

            for row_j in range(0, len(empQuery)):
                code = str(empQuery.iloc[row_j, 2])
                op = empQuery.iloc[row_j, 3]
                bundle = str(empQuery.iloc[row_j, 4])
                sumbundle = sumbundle + int(float(bundle))
                sah = str(round(empQuery.iloc[row_j, 5], 4))
                if float(sah) > float(max_sah):
                    max_sah = float(sah)
                    main_op = op
                if op == '' and bundle == '1':
                    tt = 1
                else:
                    # element = {'ID': ID, 'ID_Value': int(float(ID)), 'Name': name, 'Line': line, 'Group': group,
                    #            'Shift': shift, 'WorkHrs': workHrs, 'OT': otActual, 'OffSTD': offStd, 'Code': code,
                    #            'Operation': op, 'Bundle': bundle, 'SAH': sah, 'PlantEff': dolEff, 'Efficiency': eff}
                    element = {'ID': ID, 'ID_Value': int(float(ID)), 'Name': name, 'Line': line, 'Group': group,
                               'Shift': shift, 'WorkHrs': workHrs, 'OT': otActual, 'code02': c02, 'code04': c04,
                               'code05': c05, 'code06': c06,
                               'code07': c07, 'code09': offStd, 'code10': c10, 'code11': c11, 'code12': c12,
                               'code13': c13, 'Code': code,
                               'Operation': op, 'Bundle': bundle, 'SAH': sah, 'PlantEff': dolEff, 'Efficiency': eff}
                    dataset.append(element)

            if len(empQuery) == 1:
                if main_op != '' and main_op != None and 'None' != ID:
                    sql = "SELECT count(*) as count FROM mms.machine_operation where operation_name = '" + main_op + "' "
                    countMC = pd.read_sql(sql, engine)
                    machinecount = countMC.iloc[0, 0]
                    if machinecount > 0:
                        machine_count = '1'

        if main_op != '' and 'None' != ID:
            sah7000_query = sah7000.query('EMPLOYEE=="' + str(ID[1:6]) + '" and OPERATION=="' + str(main_op) + '"')
            if len(sah7000_query) > 0:
                is_sah7000 = 'CÓ'
        incentive = sumSAH * 7500
        income = incentive + 21933 * round(regPayroll + float(ot15) * 1.5 + float(ot20) * 2.0, 2)
        if 'M7-OT' in shift and regPayroll >= 7:
            income = incentive + 21933 * round(regPayroll + 1 + float(ot15) * 1.5 + float(ot20) * 2.0, 2)
        if cnOT == 'Y':
            if ot15 != 0 or ot20 != 0 or ot30 != 0:
                income = incentive + 21933 * round(float(ot15) * 1.5 + float(ot20) * 2.0 + float(ot30) * 3.0, 2)
        effElement2 = {'ID': ID, 'Name': name, 'Line': line, 'Group': group, 'Shift': shift, 'WorkHrs': workHrs,
                       'OT': otActual, 'OffSTD': offStd, 'PlantEff': plantEff, 'Bundle': sumbundle}
        dataset2.append(effElement2)
        effElement = {'ID': ID, 'Name': name, 'Line': line, 'Group': group, 'Shift': shift, 'WorkHrs': regPayroll,
                      'OT': otActual, 'OffSTD': round(offStd, 3), 'Incentive': incentive, 'Income': income,
                      'PlantEff': plantEff, 'DOLEff': dolEff, 'Bundle': sumbundle, 'MAIN_OPERATION': main_op,
                      'SAH7000': is_sah7000, 'DKN_OPERATION': dkn, 'tem_ot': tem_ot, 'eff_ot': eff_ot, 'MachineCount': machine_count, 'training': offStdc10}
        effDataset.append(effElement)
    dataFrame = pd.DataFrame(dataset)
    dataFrameEff = pd.DataFrame(effDataset)
    dataFrameEff2 = pd.DataFrame(dataset2)
    print(dataFrameEff)
    dataEff = dataFrameEff.sort_values(by=['Shift', 'Line', 'ID'])
    dataEff = dataFrameEff.reset_index(drop=True)
    return dataFrame, dataFrameEff2, dataEff


def getGroup():
    groupData = pd.read_sql(
        'select distinct NameGroup from setup_location where Department like "%PR" AND NameGroup IS NOT NULL AND NameGroup != ""',
        engineNam)
    engineNam.dispose()
    if len(groupData) > 0:
        return True, groupData['NameGroup'].unique()
    else:
        return False, 1


def getLine(group, dataEff):
    lineData = dataEff.query('Group=="' + group + '"')
    lineData = lineData.sort_values(by=['Line'])
    if len(lineData) > 0:
        return True, lineData['Line'].unique()
    else:
        return False, 1


def getEmplist(line, shift, dataEff):
    emplistData = dataEff.query('Line=="' + line + '" and Shift.str.contains("' + shift + '")',
                                engine='python')  # pd.read_sql('select ID, Name, Line from setup_emplist where Line="'+line+'" and Shift like "'+shift+'%"', engineNam)
    engineNam.dispose()
    if len(emplistData) > 0:
        return True, emplistData
    else:
        return False, 1


def getSupervisor(group, shift):
    svName = ''
    if shift == 'RIT':
        supervisorData = pd.read_sql('select SupervisorRitmo from setup_group where NameGroup="' + group + '";',
                                     engineNam)
        engineNam.dispose()
        if len(supervisorData) > 0:
            svName = supervisorData.iloc[0, 0]
    elif shift == 'BALI':
        supervisorData = pd.read_sql('select SupervisorBali from setup_group where NameGroup="' + group + '";',
                                     engineNam)
        engineNam.dispose()
        if len(supervisorData) > 0:
            svName = supervisorData.iloc[0, 0]
    return svName


def writeReport(shift, date, dataEff):
    thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'),
                                                 right=openpyxl.styles.borders.Side(style='thin'),
                                                 top=openpyxl.styles.borders.Side(style='thin'),
                                                 bottom=openpyxl.styles.borders.Side(style='thin'))
    fontGaramond = openpyxl.styles.Font(name='Garamond')
    if shift == 'RIT':
        wb_template.active = 0
    elif shift == 'BALI':
        wb_template.active = 1
    else:
        wb_template.active = 2
    sheet_template = wb_template.active
    header_row = 4
    row_template = 5
    rG, groupData = getGroup()
    if rG == False:
        return
    for group in groupData:
        rL, lineData = getLine(group, dataEff)
        if rL == False:
            continue
        supervisor = getSupervisor(group, shift)
        sheet_template.cell(row=header_row - 1, column=1).value = group
        sheet_template.cell(row=header_row - 2, column=6).value = shift
        sheet_template.cell(row=header_row - 2, column=8).value = date
        sheet_template.cell(row=header_row - 2, column=3).value = supervisor
        sum_reg = 0
        sum_ot = 0
        sum_09 = 0
        sum_eff = 0
        sum_hc = 0
        sum_dlo = 0
        sum_10 = 0
        count_hc = 0
        nv_2m = 0
        count_dkn = 0

        count_all = 0
        count_sah = 0
        for line in lineData:
            start_cell = row_template
            rE, emplistData = getEmplist(line, shift, dataEff)
            if rE == False:
                continue
            for row in range(0, len(emplistData)):
                ID = emplistData.iloc[row, 0]
                name = emplistData.iloc[row, 1]
                regHrs = emplistData.iloc[row, 5]
                ot = emplistData.iloc[row, 6]
                offStd = emplistData.iloc[row, 7]
                incentive = emplistData.iloc[row, 8]
                income = emplistData.iloc[row, 9]
                plantEff = emplistData.iloc[row, 10]
                dolEff = emplistData.iloc[row, 11]
                bundle = emplistData.iloc[row, 12]
                main_op = emplistData.iloc[row, 13]
                is_sah7000 = emplistData.iloc[row, 14]
                dkn = emplistData.iloc[row, 15]

                tem_ot = emplistData.iloc[row, 16]
                eff_ot = emplistData.iloc[row, 17]

                machineCount = emplistData.iloc[row, 18]
                training = emplistData.iloc[row, 19]
                # summary
                sum_reg = sum_reg + regHrs
                sum_ot = sum_ot + ot
                sum_09 = sum_09 + offStd
                sum_10 = sum_10 + training
                # count_hc = count_hc + 1
                # if (main_op != "") and ("PACKING" not in main_op) and ("CASING" not in main_op) and ("STICKER" not in main_op):
                #     count_hc = count_hc + 1
                if main_op != "":
                    if "PACKING" != main_op and "CASING" != main_op:
                        if "STICKER(1)" != main_op and "STICKER(2)" != main_op and "STICKER(3)" != main_op and "STICKER(4)" != main_op:
                            count_hc = count_hc + 1

                if machineCount == '2':
                    nv_2m = nv_2m + 1

                if dkn != '':
                    count_dkn = count_dkn + 1

                count_all = count_all + 1

                if is_sah7000 == 'CÓ':
                    count_sah = count_sah + 1

                if plantEff > 0 and plantEff < 300:
                    sum_eff = sum_eff + plantEff
                    sum_dlo = sum_dlo + dolEff
                    sum_hc = sum_hc + 1
                for col in range(1, 20):
                    sheet_template.cell(row=row_template, column=col).border = thin_border
                sheet_template.cell(row=row_template, column=1).value = emplistData.iloc[row, 2][5:8]
                sheet_template.cell(row=row_template, column=1).alignment = openpyxl.styles.Alignment(
                    horizontal='center', vertical='center')
                sheet_template.cell(row=row_template, column=2).value = ID
                sheet_template.cell(row=row_template, column=2).font = fontGaramond
                sheet_template.cell(row=row_template, column=3).value = name
                sheet_template.cell(row=row_template, column=3).font = fontGaramond
                # Add Reg/Off/Incentive/Income/Plan/Dlo/Bundle
                sheet_template.cell(row=row_template, column=4).value = regHrs  # reg
                sheet_template.cell(row=row_template, column=4).font = fontGaramond
                sheet_template.cell(row=row_template, column=5).value = ot  # over
                sheet_template.cell(row=row_template, column=5).font = fontGaramond
                sheet_template.cell(row=row_template, column=6).value = offStd  # of
                sheet_template.cell(row=row_template, column=6).font = fontGaramond

                sheet_template.cell(row=row_template, column=7).value = training  # of 10 moi them
                sheet_template.cell(row=row_template, column=7).font = fontGaramond

                sheet_template.cell(row=row_template, column=8).value = incentive  # incentive
                sheet_template.cell(row=row_template, column=8).number_format = '#,##0'
                sheet_template.cell(row=row_template, column=8).font = fontGaramond
                sheet_template.cell(row=row_template, column=9).value = income  # income
                sheet_template.cell(row=row_template, column=9).number_format = '#,##0'
                sheet_template.cell(row=row_template, column=9).font = fontGaramond
                sheet_template.cell(row=row_template, column=10).value = plantEff  # plant
                sheet_template.cell(row=row_template, column=10).font = fontGaramond
                sheet_template.cell(row=row_template, column=11).value = dolEff  # dlo
                sheet_template.cell(row=row_template, column=11).font = fontGaramond
                sheet_template.cell(row=row_template, column=12).value = bundle  # bundle
                sheet_template.cell(row=row_template, column=12).font = fontGaramond
                sheet_template.cell(row=row_template, column=12).alignment = openpyxl.styles.Alignment(
                    horizontal='center')

                sheet_template.cell(row=row_template, column=17).value = dkn  # dkn operation
                sheet_template.cell(row=row_template, column=17).font = fontGaramond

                sheet_template.cell(row=row_template, column=13).value = tem_ot  # tem ot
                sheet_template.cell(row=row_template, column=13).font = fontGaramond
                sheet_template.cell(row=row_template, column=13).alignment = openpyxl.styles.Alignment(
                    horizontal='center')

                sheet_template.cell(row=row_template, column=14).value = eff_ot  # eff ot
                sheet_template.cell(row=row_template, column=14).font = fontGaramond
                sheet_template.cell(row=row_template, column=14).alignment = openpyxl.styles.Alignment(
                    horizontal='center')

                if 'None' != str(main_op):

                    sheet_template.cell(row=row_template, column=15).value = machineCount  # số máy vận hành
                    sheet_template.cell(row=row_template, column=15).font = fontGaramond
                    sheet_template.cell(row=row_template, column=15).alignment = openpyxl.styles.Alignment(
                        horizontal='center')

                    sheet_template.cell(row=row_template, column=16).value = main_op[0:13]  # main operation
                    sheet_template.cell(row=row_template, column=16).font = fontGaramond
                    # sheet_template.cell(row=row_template, column=13).alignment=openpyxl.styles.Alignment(horizontal='center')
                    sheet_template.cell(row=row_template, column=18).value = is_sah7000  # is sah 7000
                    sheet_template.cell(row=row_template, column=18).font = fontGaramond
                    sheet_template.cell(row=row_template, column=18).alignment = openpyxl.styles.Alignment(
                        horizontal='center')
                row_template = row_template + 1
            stop_cell = row_template - 1
            if stop_cell - start_cell > 0:
                sheet_template.merge_cells('A' + str(start_cell) + ':A' + str(stop_cell))
        sheet_template.cell(row=header_row - 1, column=4).value = sum_reg
        sheet_template.cell(row=header_row - 1, column=4).font = fontGaramond
        sheet_template.cell(row=header_row - 1, column=5).value = sum_ot
        sheet_template.cell(row=header_row - 1, column=5).font = fontGaramond
        sheet_template.cell(row=header_row - 1, column=6).value = sum_09
        sheet_template.cell(row=header_row - 1, column=6).font = fontGaramond
        sheet_template.cell(row=header_row - 1, column=7).value = sum_10
        sheet_template.cell(row=header_row - 1, column=7).font = fontGaramond

        if count_hc == 0:
            value15 = 0
        else:
            value15 = (nv_2m / count_hc) * 100
        sheet_template.cell(row=header_row - 1, column=15).value = round(value15, 2)
        sheet_template.cell(row=header_row - 1, column=15).font = fontGaramond

        if count_hc == 0:
            value17 = 0
        else:
            value17 = (count_dkn / count_hc) * 100
        sheet_template.cell(row=header_row - 1, column=17).value = round(value17, 2)
        sheet_template.cell(row=header_row - 1, column=17).font = fontGaramond

        if count_all == 0:
            value18 = 0
        else:
            value18 = (count_sah / count_all) * 100
        sheet_template.cell(row=header_row - 1, column=18).value = round(value18, 2)
        sheet_template.cell(row=header_row - 1, column=18).font = fontGaramond

        if sum_hc != 0:
            sheet_template.cell(row=header_row - 1, column=10).value = round(sum_eff / sum_hc, 2)
        else:
            sheet_template.cell(row=header_row - 1, column=10).value = 0
        sheet_template.cell(row=header_row - 1, column=10).font = fontGaramond
        # sheet_template.cell(row=header_row-1, column=10).value=sum_dlo
        # sheet_template.cell(row=header_row-1, column=10).font=fontGaramond
        row_template = row_template + 1
        for row in range(1, 5):
            try:
                for col in range(1, 20):
                    sheet_template.cell(row=row_template, column=col).value = sheet_template.cell(row=row,
                                                                                                  column=col).value
                    if sheet_template.cell(row=row, column=col).has_style:
                        sheet_template.cell(row=row_template, column=col)._style = sheet_template.cell(row=row,
                                                                                                       column=col)._style
                row_template = row_template + 1
            except:
                tt = 1
        sheet_template.merge_cells('C' + str(row_template - 4) + ':O' + str(row_template - 4))
        header_row = row_template - 1


def get_offstd_time(ID, shift, dateFull):
    # off standard
    cd03 = 0
    cd08 = 0
    cd09 = 0
    cd15 = 0
    offStdQuery = pd.read_sql(
        'select Code, SUM(SpanTime) as OffTime from operation_offstandard_tracking where right(ID,5) = right("' + ID + '",5) and DateUpdate="' + dateFull + ' 00:00:00" and IEApprovedResult="2" GROUP by LEFT(Code, 2);',
        engineLine)
    engineLine.dispose()
    # if ID==''
    if len(offStdQuery) > 0:
        for row in range(0, len(offStdQuery)):
            offStd = 0
            try:
                offStd = float(offStdQuery.iloc[row, 1])
                if offStd > 7.37:
                    offStd = 7.37
                else:
                    if ('CHILD' in shift or 'MAT' in shift) and offStd > 6.37:
                        offStd = 6.37
            except:
                offStd = 0
            if '03' in offStdQuery.iloc[row, 0]:
                cd03 = offStd
            elif '08' in offStdQuery.iloc[row, 0]:
                cd08 = offStd
            elif '09' in offStdQuery.iloc[row, 0]:
                cd09 = offStd
            # elif '15' in offStdQuery.iloc[row, 0]:
            #     cd15=offStd
    if cd03 < 0:
        cd03 = 0
    if cd08 < 0:
        cd08 = 0
    if cd09 < 0:
        cd09 = 0
    # if cd15<0:
    #     cd15=0
    return cd03, cd08, cd09  # , cd15


def get_offstd_time_cutting(ID, shift, dateFull):
    # off standard
    cd03 = 0
    cd08 = 0
    cd09 = 0
    offStdQuery = pd.read_sql(
        'select OffCode, SUM(Duration) as OffTime from cutting_system.offstandard_employee_tracking where ID = "' + ID + '" and DATE_IN="' + dateFull + '" and Approved like "2%" GROUP by LEFT(OffCode, 2);',
        engineLine)
    # print('select Code, SUM(Duration) as OffTime from cutting_system.offstandard_employee_tracking where ID = "'+ID+'" and DATE_IN="'+dateFull+'" and Approved like "2%" GROUP by LEFT(OffCode, 2);')
    engineLine.dispose()
    # if ID==''
    if len(offStdQuery) > 0:
        for row in range(0, len(offStdQuery)):
            offStd = 0
            try:
                offStd = float(offStdQuery.iloc[row, 1])
                if offStd > 7.37:
                    offStd = 7.37
                else:
                    if ('CHILD' in shift or 'MAT' in shift) and offStd > 6.37:
                        offStd = 6.37
            except:
                offStd = 0
            if '03' in offStdQuery.iloc[row, 0]:
                cd03 = offStd
            elif '08' in offStdQuery.iloc[row, 0]:
                cd08 = offStd
            elif '09' in offStdQuery.iloc[row, 0]:
                cd09 = offStd
    if cd03 < 0:
        cd03 = 0
    if cd08 < 0:
        cd08 = 0
    if cd09 < 0:
        cd09 = 0
    return cd03, cd08, cd09


def get_whr_by_date(date):
    rit = 8
    bal = 8
    rit_ot = 'N'
    bal_ot = 'N'
    return rit, bal, rit_ot, bal_ot


def send_mail(link1, link2, date):
    html = '<!DOCTYPE html><html>'
    html = html + "<h4>Xuất báo cáo thành công</h4><p>Báo cáo lương sản phẩm: " + link1 + "</p><p>Báo cáo hiệu suất: " + link2 + "</p><p>Thanks, PBAssistant</p>"
    username = "PBAssistant@hanes.com"
    password = "xLkUmsMZkbPaLAwcWSgMGhwWMTYk67"
    address_book = ['HBI_Phubai_Engineering@hanes.com', 'HBI_Phubai_Operations_Management@hanes.com']
    # address_book = ['dat.la@hanes.com']
    msg = MIMEMultipart()
    msg['From'] = username
    msg['To'] = ','.join(address_book)
    msg['Subject'] = "Tự động _ Báo cáo lương sản phẩm và hiệu suất ngày " + date
    msg.attach(MIMEText(html, 'html'))
    try:
        mailServer = smtplib.SMTP('smtp-mail.outlook.com', 587)
        mailServer.ehlo()
        mailServer.starttls()
        mailServer.ehlo()
        mailServer.login(username, password)
        mailServer.send_message(msg)
        mailServer.close()
        print('Email sent successfully')
    except:
        print('An error has occurred when send mail')


def get_offStd_all(date, employee):
    data = pd.read_sql(
        'SELECT SpanTime, left(CODE,2) as offstd FROM operation_offstandard_tracking WHERE DateUpdate = "' + date + ' 00:00:00" AND right(ID,5) = right("' + employee + '",5) and (IEApprovedResult="2" or (IEApprovedResult="0" AND left(CODE,2) = "05"));',
        engineLine)

    code02 = data.query('offstd=="02"')
    if len(code02) == 0:
        c02 = 0
    else:
        c02 = code02.iloc[0, 0]
    code04 = data.query('offstd=="04"')
    if len(code04) == 0:
        c04 = 0
    else:
        c04 = code04.iloc[0, 0]

    code05 = data.query('offstd=="05"')
    if len(code05) == 0:
        c05 = 0
    else:
        c05 = code05.iloc[0, 0]

    code06 = data.query('offstd=="06"')
    if len(code06) == 0:
        c06 = 0
    else:
        c06 = code06.iloc[0, 0]

    code07 = data.query('offstd=="07"')
    if len(code07) == 0:
        c07 = 0
    else:
        c07 = code07.iloc[0, 0]

    code10 = data.query('offstd=="10"')
    if len(code10) == 0:
        c10 = 0
    else:
        c10 = code10.iloc[0, 0]

    code11 = data.query('offstd=="11"')
    if len(code11) == 0:
        c11 = 0
    else:
        c11 = code11.iloc[0, 0]

    code12 = data.query('offstd=="12"')
    if len(code12) == 0:
        c12 = 0
    else:
        c12 = code12.iloc[0, 0]

    code13 = data.query('offstd=="13"')
    if len(code13) == 0:
        c13 = 0
    else:
        c13 = code13.iloc[0, 0]

    return c02, c04, c05, c06, c07, c10, c11, c12, c13

def get_shift_by_date(d):
    date_str = d.replace("-", "")
    # date_str = d.strftime('%Y%m%d')
    sql = ("SELECT * FROM operation_schedule o WHERE o.DATE = '"+date_str+"';")
    shift_morning = pd.read_sql(sql, engine2)
    sm = shift_morning['Shift'][0]
    if sm == 'R':
        sf = 'B'
    else:
        sf = 'R'
    return sm, sf

def get_tem_ot(employee, date):
    morning, afternoon = get_shift_by_date(date)
    date_str = date.replace("-", "")
    sql = ("SELECT shift FROM erpsystem.setup_emplist where id = '"+employee+"';")
    data = pd.read_sql(sql, engine2)
    shift = data.iloc[0, 0]
    r5 = employee[-5:]

    if 'BALI' in shift:
        if morning == 'R':
            query = ("SELECT ticket, earned_hours, TimeUpdate FROM pr2k.employee_scanticket a WHERE a.DATE = '"+date_str+"' AND a.EMPLOYEE = '"+r5+"' "
                     "AND a.EARNED_HOURS != 0 AND a.TimeUpdate < '"+date+" 14:20:00';")
            data = pd.read_sql(query, engine2)
            return data
        else:
            query = ("SELECT ticket, earned_hours, TimeUpdate FROM pr2k.employee_scanticket a WHERE a.DATE = '" + date_str + "' AND a.EMPLOYEE = '" + r5 + "' "
                     "AND a.EARNED_HOURS != 0 AND a.TimeUpdate > '" + date + " 14:20:00';")
            data = pd.read_sql(query, engine2)
            return data
    if 'RIT' in shift:
        if morning == 'R':
            query = ("SELECT ticket, earned_hours, TimeUpdate FROM pr2k.employee_scanticket a WHERE a.DATE = '" + date_str + "' AND a.EMPLOYEE = '"+r5+"' "
                        "AND a.EARNED_HOURS != 0 AND a.TimeUpdate > '" + date + " 14:20:00';")
            data = pd.read_sql(query, engine2)
            return data
        else:
            query = ("SELECT ticket, earned_hours, TimeUpdate FROM pr2k.employee_scanticket a WHERE a.DATE = '" + date_str + "' AND a.EMPLOYEE = '" + r5 + "' "
                      "AND a.EARNED_HOURS != 0 AND a.TimeUpdate < '" + date + " 14:20:00';")
            data = pd.read_sql(query, engine2)
            return data
    return pd.DataFrame([])

if __name__=="__main__":
    today=datetime.today()
    #get yesterday
    # yesterday=today-timedelta(days=1)
    yesterday=date(2024,10,28)
    thisWeek=get_this_week(yesterday)
    dailyEfficiency_link='\\\\pbvfps1\\Bundle_Scan\\Report\\DailyEfficiencyReport\\'
    yesterday_str=yesterday.strftime('%Y%m%d')
    yesterday_str2=yesterday.strftime('%Y-%m-%d')
    fnameFile=dailyEfficiency_link+'\\'+thisWeek+'\\DailyEfficiencyReportt_'+yesterday_str+'.xlsx'
    # print(yesterday_str)
    rit_whr, bal_whr, rit_ot, bal_ot = get_whr_by_date(yesterday_str)
    print('work hour = ')
    print(rit_whr)
    result = 0
    #upload timesheet to database
    # if True:
    # is_exist=False
    # try:
    process_timesheet(yesterday_str2, rit_whr, bal_whr, rit_ot, bal_ot)
    # except:
        # print('fail-TMS_'+yesterday_str)
    #     quit()
    # get Daily Efficiency of yesterday
    # if os.path.exists(fnameFile):
    #     is_exist=True
    # try:
    if True:
        dataFrame, dataFrameEff2, dataEff=getEff(yesterday_str)
        if len(dataEff)>0:
            #save day Daily Efficiency to specific link
            dailyEfficiency_link='\\\\pbvfps1\\Bundle_Scan\\Report\\DailyEfficiency\\'
            thisWeek=get_this_week(yesterday)
            if not os.path.exists(dailyEfficiency_link+thisWeek):
                os.makedirs(dailyEfficiency_link+'\\'+thisWeek)
            writer = pd.ExcelWriter(dailyEfficiency_link+'\\'+thisWeek+'\\DailyEfficiencyy_'+yesterday_str+'.xlsx', engine='xlsxwriter')
            dataFrame.to_excel(writer, sheet_name='DetailEfficiency', index=False)
            dataFrameEff2.to_excel(writer, sheet_name='DailyEfficiency', index=False)
            try:
                writer.save()
            except:
                result=1
            #run yesterday
            try:
                writeReport('RIT', yesterday_str, dataEff)
                writeReport('BALI', yesterday_str, dataEff)
                writeReport('AD', yesterday_str, dataEff)
            except:
                result=2
            fpath=''
            dailyEfficiency_link='\\\\pbvfps1\\Bundle_Scan\\Report\\DailyEfficiencyReport\\'
            if not os.path.exists(dailyEfficiency_link+thisWeek):
                os.makedirs(dailyEfficiency_link+thisWeek)
            try:
                wb_template.save(fnameFile)
            except:
                result=3
            wb_template.close()
            result=4
        else:
            result=5
    # except:
        print('fail-'+yesterday_str)
        link1 = '\\\\pbvfps1\\Bundle_Scan\\Report\\DailyEfficiencyReport\\' + thisWeek + '\\DailyEfficiencyReport_' + yesterday_str + '.xlsx'
        link2 = '\\\\pbvfps1\\Bundle_Scan\\Report\\DailyEfficiency\\' + thisWeek + '\\DailyEfficiency_' + yesterday_str + '.xlsx'
        # send_mail(link1, link2, yesterday_str)
    else:
        # if is_exist==True:
        print('done-'+yesterday_str)
